import os
import re
import pdfplumber
import pandas as pd
import pdfkit
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import babel.numbers

import locale

locale.setlocale(locale.LC_ALL, "pt_BR.UTF-8")  # Set Brazilian currency format

# Configuration constants
INPUT_FILES_DIR = "./input"
OUTPUT_DIR = "./output"
CORRECT_TAX_RATE = 0.1508

# Create output directory if it doesn't exist
os.makedirs(OUTPUT_DIR, exist_ok=True)


def style_excel_worksheet(ws):
    """Apply professional styling to Excel worksheet."""
    # Define colors (matching HTML style)
    header_color = "000000"  # Black
    alternate_row_color = "F2F2F2"  # Light gray
    border_color = "000000"  # Black text

    # Define styles
    header_fill = PatternFill(
        start_color=header_color, end_color=header_color, fill_type="solid"
    )
    alternate_fill = PatternFill(
        start_color=alternate_row_color,
        end_color=alternate_row_color,
        fill_type="solid",
    )
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Arial", size=10, color="000000")

    border = Border(
        left=Side(style="thin", color=border_color),
        right=Side(style="thin", color=border_color),
        top=Side(style="thin", color=border_color),
        bottom=Side(style="thin", color=border_color),
    )

    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Style data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            if row_idx % 2 == 0:  # Apply alternate row color
                cell.fill = alternate_fill

    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width


def format_brl(value):
    """Convert a float to Brazilian Real currency format (R$ 1.234,56)."""
    if pd.isna(value) or value == "":
        return "R$ 0,00"
    return babel.numbers.format_currency(value, "BRL", locale="pt_BR")


def generate_reports(df):
    """Generate both Excel and PDF reports with correct BRL formatting."""

    # Rename columns to Portuguese
    portuguese_columns = {
        "client_number": "Nº do cliente",
        "installation_number": "Nº do contrato",
        "calculation_base": "Base de Cálculo",
        "applied_tax_rate": "Taxa de ICMS Aplicada",
        "fixed_tax_rate": "Taxa de ICMS Corrigida",
        "paid_icms": "Valor de ICMS pago",
        "correct_icms": "Valor de ICMS corrigido",
        "credits": "Créditos",
        "period": "Período",
    }
    df = df.rename(columns=portuguese_columns)

    # Ensure 'Créditos' column exists
    if "Créditos" not in df.columns:
        df["Créditos"] = 0

    # Apply BRL formatting
    for col in [
        "Base de Cálculo",
        "Valor de ICMS pago",
        "Valor de ICMS corrigido",
        "Créditos",
    ]:
        df[col] = df[col].apply(format_brl)

    # Properly convert 'Créditos' column for summing
    total_credits = (
        df["Créditos"]
        .apply(
            lambda x: float(
                x.replace("R$", "")
                .replace("\xa0", "")
                .replace(".", "")
                .replace(",", ".")
            )
        )
        .sum()
    )

    # Generate HTML for PDF
    html_content = f"""
    <html>
    <head>
        <link href="https://fonts.googleapis.com/css2?family=Manrope:wght@300;400;700&display=swap" rel="stylesheet">
        <style>
            body {{ font-family: 'Manrope', sans-serif; margin: 20px; font-size: 10px; color: black; background-color: white; }}
            h1 {{ text-align: center; font-size: 16px; font-weight: bold; margin-bottom: 10px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
            th {{ background-color: black; color: white; padding: 8px; text-align: center; font-size: 10px; }}
            td {{ padding: 6px; text-align: center; border: 1px solid black; font-size: 10px; }}
            tr:nth-child(even) {{ background-color: #f2f2f2; }}
            .total-row {{ font-weight: bold; background-color: white; border: none; text-align: right; }}
            .total-row td {{ border: none; font-size: 12px; }}
        </style>
    </head>
    <body>
        <h1>Relatório de Análise ICMS</h1>
        {df.to_html(index=False, classes='styled-table', escape=False)}
        
        <table>
            <tr class="total-row">
                <td colspan="{len(df.columns) - 1}"><strong>Total À Recuperar</strong></td>
                <td><strong>{format_brl(total_credits)}</strong></td>
            </tr>
        </table>
    </body>
    </html>
    """

    pdf_path = os.path.join(OUTPUT_DIR, "relatorio_icms.pdf")
    pdfkit.from_string(
        html_content,
        pdf_path,
        options={
            "page-size": "A4",
            "margin-top": "20mm",
            "margin-right": "20mm",
            "margin-bottom": "20mm",
            "margin-left": "20mm",
            "encoding": "UTF-8",
            "no-outline": None,
        },
    )

    return os.path.join(OUTPUT_DIR, "relatorio_icms.xlsx"), pdf_path


def extract_client_and_installation(text: str) -> tuple[str | None, str | None]:
    """Extract client and installation numbers from the PDF text using regex."""
    client_match = re.search(r"Nº DO CLIENTE\n(\d+)", text)
    installation_match = re.search(r"Nº DA INSTALAÇÃO\n(\d+)", text)

    client_number = client_match.group(1) if client_match else None
    installation_number = installation_match.group(1) if installation_match else None

    return client_number, installation_number


def parse_brazilian_float(value_str: str) -> float | None:
    """Convert Brazilian number format to float."""
    clean_string = value_str.replace(".", "").replace(",", ".")
    try:
        return float(clean_string)
    except ValueError:
        return None


def _process_tax_info_layout(tables: list, text_lines: list) -> None:
    """Process pages with 'INFORMAÇÕES DE TRIBUTOS' layout."""
    for table in tables:
        if len(table) == 15:
            client_number, installation_number = extract_client_and_installation(
                table[2][-1]
            )
            calculation_base = parse_brazilian_float(table[-1][0])
            applied_tax_rate = parse_brazilian_float(table[-1][1])
            paid_icms = parse_brazilian_float(table[-1][2])

            _extract_and_add_tax_record(
                text_lines,
                client_number,
                installation_number,
                calculation_base,
                applied_tax_rate,
                paid_icms,
            )


def _process_standard_layout(tables: list, text_lines: list) -> None:
    """Process pages with standard layout."""
    client_number = installation_number = current_month = None

    for table in tables:
        if len(table) == 2 and "CÓDIGO DA INSTALAÇÃO" in table[0][0]:
            client_number = table[0][0].split("\n")[1]
            installation_number = table[1][0].split("\n")[1]

    for i, line in enumerate(text_lines):
        words = line.split()

        if (
            i > 0
            and "TOTAL A PAGAR" in text_lines[i - 1]
            and "REF:MÊS/ANO" in text_lines[i - 1]
        ):
            current_month = words[0]

        if len(words) >= 4:
            _process_icms_line(
                words[-4:], client_number, installation_number, current_month
            )


def _process_icms_line(
    values: list, client_number: str, installation_number: str, current_month: str
) -> None:
    """Process a line containing ICMS tax information."""
    if values[0] != "ICMS":
        return

    try:
        if not all(all(c.isdigit() or c in ",." for c in num) for num in values[1:]):
            print("Invalid number format found:", values[1:])
            return

        calculation_base = parse_brazilian_float(values[1])
        applied_tax_rate = parse_brazilian_float(values[2])
        paid_icms = parse_brazilian_float(values[3])

        if all(v is not None for v in [calculation_base, applied_tax_rate, paid_icms]):
            _add_tax_record(
                client_number,
                installation_number,
                current_month,
                calculation_base,
                applied_tax_rate,
                paid_icms,
            )
    except Exception as e:
        print(f"Error processing ICMS line: {e}")


def _add_tax_record(
    client_number: str,
    installation_number: str,
    period: str,
    calculation_base: float,
    applied_tax_rate: float,
    paid_icms: float,
) -> None:
    """Add a new tax record to the DataFrame."""
    correct_icms = round(CORRECT_TAX_RATE * calculation_base, 2)
    tax_credits = round(paid_icms - correct_icms, 2)

    tax_data_df.loc[len(tax_data_df)] = {
        "client_number": client_number,
        "installation_number": installation_number,
        "period": period,
        "calculation_base": calculation_base,
        "applied_tax_rate": f"{round(applied_tax_rate, 2)}%",
        "fixed_tax_rate": f"{round(CORRECT_TAX_RATE * 100, 2)}%",
        "correct_icms": correct_icms,
        "paid_icms": paid_icms,
        "credits": tax_credits,
    }


def _extract_and_add_tax_record(
    text_lines: list,
    client_number: str,
    installation_number: str,
    calculation_base: float,
    applied_tax_rate: float,
    paid_icms: float,
) -> None:
    """Extract period information and add tax record for tax info layout."""
    for i, line in enumerate(text_lines):
        if "CONTA CONTRATO" in text_lines[i - 1]:
            words = line.split()
            if len(words) > 1 and "/" in words[1]:
                _add_tax_record(
                    client_number,
                    installation_number,
                    words[1],
                    calculation_base,
                    applied_tax_rate,
                    paid_icms,
                )


def process_tax_page(tables: list, text: str, page_num: int) -> None:
    """Process a single page of tax information from the PDF."""
    print(f"------------------\nProcessing PAGE #{page_num}---------------\n")

    text_lines = text.splitlines()
    if "INFORMAÇÕES DE TRIBUTOS" in text:
        _process_tax_info_layout(tables, text_lines)
    else:
        _process_standard_layout(tables, text_lines)


if __name__ == "__main__":
    # Initialize DataFrame
    tax_data_columns = {
        "installation_number": [],
        "client_number": [],
        "period": [],
        "calculation_base": [],
        "applied_tax_rate": [],
        "paid_icms": [],
        "fixed_tax_rate": [],
        "correct_icms": [],
        "credits": [],
    }
    tax_data_df = pd.DataFrame(tax_data_columns)

    # Get list of files to process
    input_files = os.listdir(INPUT_FILES_DIR)
    print("Files to process:", input_files)

    # Process each PDF file
    for file in input_files:
        if not file.endswith(".pdf"):
            continue

        file_path = os.path.join(INPUT_FILES_DIR, file)
        print("Processing file:", file)

        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text()
                tables = page.extract_tables()
                process_tax_page(tables, text, page_num)

    # Prepare final report
    tax_data_df["period"] = pd.to_datetime(
        tax_data_df["period"], format="%m/%Y"
    ).dt.strftime("%m/%Y")
    tax_data_df = tax_data_df.sort_values(by="period")

    # Generate reports
    excel_path, pdf_path = generate_reports(tax_data_df)
    print(f"\nReports generated successfully:")
    print(f"Excel report: {excel_path}")
    print(f"PDF report: {pdf_path}")
